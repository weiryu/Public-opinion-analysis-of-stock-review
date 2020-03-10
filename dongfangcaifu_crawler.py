# -*- coding:UTF-8 -*-
"""
# 爬取东方财富股吧发帖评论数据
"""

import sys
import importlib
import time
import re, requests, codecs, time, random
import pandas as pd
from lxml import html
from bs4 import BeautifulSoup
import openpyxl
import os
import random


#复制请求头
head ={
'Accept-Encoding':'gzip,deflate',
'Accept-Language':'zh-CN,zh;q=0.9',
'Connection':'keep-alive',
'Host':'guba.eastmoney.com',
'User-Agent':'Mozilla/5.0(WindowsNT6.1;Win64;x64)AppleWebKit/537.36(KHTML,likeGecko)Chrome/65.0.3325.181Safari/537.36'}

# 正则表达式获得页面评论时间
def get_comment_time(q):
    pub_time = 0000
    try:
        #<div class="zwfbtime">发表于 2020-02-11 09:54:48 东方财富Android版</div>
        temp_1 = re.search('<div class="zwfbtime">.*?</div>',q.text).group(0)
        # 2020-02-06
        pub_time = re.search('\d\d\d\d-\d\d-\d\d',temp_1).group(0)
        # 获得发帖时间
        pub_time = int(pub_time.replace("-",''))
        # print(pub_time)
    except:
        pass
    return pub_time

# 正则表达式获得页面评论标题
def get_comment_title(q):
    title = ''
    try:
        # <title>虽然指数跌的多但主要是石油板块带指数个股跌幅不大[微笑][微笑][微笑]关注基建_上证指数(zssh000001)股吧_东方财富网股吧</title>
        temp_2 = re.search('<title>.*?</title>',q.text).group(0)
        # 获得帖子标题
        title = temp_2[7:][:-35]
        # print(title)
    except:
        pass
    return title

# 正则表达式获得页面评论用户名
def get_user_name(q):
    user_name = ''
    try: 
        temp_3 = re.search('<font>.*?</font>',q.text).group(0)
        # 获得发贴用户
        user_name = temp_3[6:][:-7]
    except:
        pass
    return user_name

# 获得评论数据
def get_comments(base_url, url, stocknum):
    q_1 = requests.get(url,headers=head)
    pattern_1 = re.compile('/news\S+html',re.S)
    news_comment_urls = re.findall(pattern_1, q_1.text) # 非空白字符N次
    comments = []
    for i in range(len(news_comment_urls)):
    	# 抽样比例为10%，只选取10%的帖子内容进行保存
    	# 可以根据服务器配置自由选取保存比例。
        if i%10 == 0:
            news_comment_url = news_comment_urls[i]
            # 合成梯子网址
            news_comment_whole_url = base_url+news_comment_url
            # 只处理与指定股票相关的帖子
            if (stocknum in (news_comment_whole_url)):
                # print(news_comment_whole_url)
                try:
                    q_2 = requests.get(news_comment_whole_url,headers=head,timeout=5)
                    # 获得发帖用户名
                    user_name = get_user_name(q_2)
                    # 获得发帖时间
                    pub_time = get_comment_time(q_2)
                    # 获得帖子标题
                    title = get_comment_title(q_2)
                    # 返回获得的帖子数据： 发贴用户+发帖时间+帖子标题+帖子链接
                    single_comment = [user_name, pub_time, title, news_comment_whole_url]
                    comments.append(single_comment)
                # 异常处理，此处忽略。异常数据直接丢弃
                except:
                    pass
            else:
                pass
    # 按发帖时间排序，一页帖子通常包含80个帖子，根据保留比例不同最终返回的帖子数量在0-80之间
    comments.sort(key=lambda x:x[1],reverse=True)
    return comments
    

def write_to_file(comments,file_path):
    # excel文件不存在，重新创建
    if not os.path.exists(file_path):
        data = openpyxl.Workbook()
        table = data.create_sheet('Sheet1')
        table = data.active
        nrows = 1
    # excel文件已经存在，打开并添加到最后
    else:
        data = openpyxl.load_workbook(file_path)
        # print(data.get_named_ranges()) # 输出工作页索引范围
        # print(data.get_sheet_names()) # 输出所有工作页的名称
        # 取第一张表
        sheetnames = data.get_sheet_names()
        table = data.get_sheet_by_name(sheetnames[0])
        table = data.active
        nrows = table.max_row # 获得行数
    # print(nrows)    
    for i in range(len(comments)):
        comment = comments[i]
        for j in range(4):
            table.cell(row = int(nrows+i+1), column = j+1).value = comment[j]
    data.save(file_path)
    time.sleep(10)

def main():
    # 爬取股吧对应的股票代码，此处选择上证综指
    stocknum = r'zssh000001'
    # 爬取股评的起始页数
    start_page_index = 1000
    # 爬取股评的终止页数
    end_page_index = 2000
    # 最长休眠时间
    max_sleep_second = 3
    # 爬取网址
    base_url = r'http://guba.eastmoney.com'
    # 内容保存文件夹
    file_folder = r''
    # 每个excel文件保存多少页评论
    pages_per_file = 10
    if not os.path.exists(file_folder):
        os.mkdir(file_folder)
    for page_index in range(start_page_index, end_page_index):
        print('Crawling to page {}'.format(page_index))
        # 构建股评帖子网址，按照发帖时间排序
        url = 'http://guba.eastmoney.com/list,' + str(stocknum) + ',f_' + str(page_index) + '.html'
        # 处理单页中的所有股评数据，按照时间排序
        comments = get_comments(base_url, url, stocknum)  
        print('Page {} is crawled,totally {} comments.'.format(page_index, len(comments)))
        file_path = os.path.join(file_folder,str(page_index//pages_per_file)+'.xlsx')
        write_to_file(comments,file_path)
        # 通过random.randint()函数返回指定闭区间内的随机数，随机休眠以便反爬
        randomInt = random.randint(1,max_sleep_second)
        print("将要阻塞 " + str(randomInt) + " 秒")
        time.sleep(randomInt)


if __name__ == '__main__':
    main()