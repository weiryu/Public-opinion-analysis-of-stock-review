"""
Analysis of emotional tendency
调用百度NLP接口实现情感倾向分析
"""

from aip import AipNlp
import openpyxl
import os
import time

# baidu api information
APP_ID = ''
API_KEY = ''
SECRET_KEY = ''
client = AipNlp(APP_ID, API_KEY, SECRET_KEY)

# set sleep time to use baidu api
def sleeptime(hour,min,sec):
    return hour*3600 + min*60 + sec
		
# 调用百度NLP的评论情感分析API，获得情感极性分类结果
def get_sentiments(text):
	result = []
	try:
		sentiment_result = client.sentimentClassify(text)['items'][0]
		positive_prob = sentiment_result['positive_prob'] #表示属于积极类别的概率
		negative_prob = sentiment_result['negative_prob'] #表示属于消极类别的概率
		confidence = sentiment_result['confidence'] #表示分类的置信度
		sentiment = sentiment_result['sentiment'] #表示情感极性分类结果, 0:负向，1:中性，2:正向
		result = [positive_prob,negative_prob,confidence,sentiment]
	# 异常处理，此处忽略
	except Exception as e:
		pass
	return result

# 对excel中的评论逐条处理并将获得的情感分析结果保存到原文件中
def process_excel(file_path):
	data = openpyxl.load_workbook(file_path)
	table = data.active
	nrows = table.max_row
	for i in range(2,nrows+1):
		# 拼接文本位置
		text_location = 'C'+str(i)
		# 提取文本内容
		text = table[text_location].value
		if text:
			# 百度评论分类
			result = get_sentiments(text)
			# 保存结果
			if result:
				table.cell(row=int(i), column=5).value = result[0]
				table.cell(row=int(i), column=6).value = result[1]
				table.cell(row=int(i), column=7).value = result[2]
				table.cell(row=int(i), column=8).value = result[3]
	data.save(file_path)


def main():
	second = sleeptime(0,0,2)
	filefolder = r''
	for file in os.listdir(filefolder):
		# only process excel file
		print(os.path.splitext(file)[-1])
		if os.path.splitext(file)[-1] in ['.xlsx']:
		    file_path = os.path.join(filefolder, file)
		    process_excel(file_path)
		    time.sleep(second) #随机休眠
		    

if __name__ == '__main__':
	main()