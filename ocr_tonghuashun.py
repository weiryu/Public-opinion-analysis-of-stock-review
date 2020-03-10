# -*- coding: utf-8 -*-
"""
对在模拟器中获得的同花顺股票评论截图
通过百度ocr进行文字识别
获得结果写入excel
"""

from aip import AipOcr
import re
import os
from PIL import ImageGrab
import time
import webbrowser
from urllib.parse import quote
from bs4 import BeautifulSoup
#从selenium里面导入webdriver
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions 
import os
import win32com
from win32com.client import constants as c  # 旨在直接使用VBA常数
import openpyxl


# baidu api information
APP_ID = ''
API_KEY = ''
SECRET_KEY = ''
client = AipOcr(APP_ID,API_KEY,SECRET_KEY)

# set sleep time to use baidu api
def sleeptime(hour,min,sec):
    return hour*3600 + min*60 + sec

# get file content
def get_file_content(filePath):
    with open(filePath,'rb') as fp:
        return fp.read()

#、process ocr result to get raw text
def raw_text_process(raw_text):
    # print(raw_text)
    temp_texts = []
    for text in str(raw_text).split(",")[2:]:
        if not re.findall("\d{6,}",text):
            text = text.replace("words_result","").replace("words","")
            temp_texts.append(re.sub(r'[\'\s\[\],:{}]*','',text))
    return temp_texts

# get ocr result
def img_to_str(filePath):
    image = get_file_content(filePath)
    print("start")
    raw_text = client.basicGeneral(image)
    result = raw_text_process(raw_text)
    return result

# split the raw text to get review one person by one person
def text_process(text):
    # text = ['s回@', '令M日733', '长江电力', '17.48-1.35%', '盘口资金论股公告简况(F10', '最新', '最热', '同顺号', '问董秘', '分享', '评论', 'mo*717b8d◎', '大盘涨你不涨', '大盘跌你跟着跌', '分享', '评论', 'johnsonbaao', '真是怂。。。垃圾白马', '分享', '四评论', '子弹飞牛氓', '过来避险啊', '分享', '评论', '股海老牛777c', 'Android发布于03-060845', '个', '场外资金很多', ' 轻仓空仓的比比皆是', '还有大量的海外资本不断', '∠我也来说两句', '发送', '电力', '0.11%', '下单', '诊股', '加自选']
    index_1 = []
    index_2 = []
    index_3 = []
    start_index = 0
    end_index = len(text)-1 
    for i,v in enumerate(text):
        Flag = False
        if not Flag and '分享' in v:
            index_1.append(i)
            Flag = True
        if not Flag and '评论' in v :
            index_2.append(i)
            Flag = True
        if not Flag and '赞'==v :
            index_3.append(i)
            Flag = True
        if not Flag and '董秘' in v:
            start_index = i
        if not Flag and '说两句' in v:
            end_index = i
    stop_index = []
    while (index_1 or index_2 or index_3):
        #print(index_1)
        #print(index_2)
        #print(index_3)
        #print(stop_index)
        if index_1: 
            if index_2:
                # index_1+index_2+index_3
                if index_3:
                    # [分享,评论,赞]
                    if (index_3[0]-index_2[0])==1 and (index_3[0]-index_1[0])==2:
                        stop_index.append(index_3[0])
                        index_1.pop(0)
                        index_2.pop(0)
                        index_3.pop(0)
                    # [分享, ,赞]
                    elif index_3[0]-index_1[0]==1 and index_2[0]>index_3[0]:
                        stop_index.append(index_3[0])
                        index_1.pop(0)
                        index_3.pop(0)
                    # [ ,评论,赞]
                    elif index_3[0]-index_2[0]==1 and index_1[0]>index_3[0]:
                        stop_index.append(index_3[0])
                        index_2.pop(0)
                        index_3.pop(0)    
                    # [分享,评论, ]    
                    elif index_3[0]-index_2[0]>1 and (index_2[0]-index_1[0])==1:
                        stop_index.append(index_2[0])
                        index_1.pop(0)
                        index_2.pop(0)
                    # [分享, , ]
                    elif index_3[0]-index_1[0]>2 and index_2[0]-index_1[0]>1:
                        stop_index.append(index_1[0])
                        index_1.pop(0)
                    # [,评论, ]
                    elif index_3[0]-index_2[0]>1 and index_2[0]<index_1[0]:
                        stop_index.append(index_2[0])
                        index_2.pop(0)
                    # [, ,赞]
                    elif index_3[0]<index_1[0] and index_3[0]<index_2[0]:
                        stop_index.append(index_3[0])
                        index_3.pop(0)    
                    else:
                        pass
                # index_1+index_2
                else:
                    # [分享,评论, ]    
                    if (index_2[0]-index_1[0])==1:
                        stop_index.append(index_2[0])
                        index_1.pop(0)
                        index_2.pop(0)
                    # [分享, , ]
                    elif index_2[0]-index_1[0]>1:
                        stop_index.append(index_1[0])
                        index_1.pop(0)
                    # [,评论, ]
                    elif index_2[0]<index_1[0]:
                        stop_index.append(index_2[0])
                        index_2.pop(0)
                    else:
                        pass
            else:
                # index_1+index_3
                if index_3:
                    # [分享, ,赞]
                    if index_3[0]-index_1[0]==1 :
                        stop_index.append(index_3[0])
                        index_1.pop(0)
                        index_3.pop(0)
                    # [分享, , ]
                    elif index_3[0]-index_1[0]>2:
                        stop_index.append(index_1[0])
                        index_1.pop(0)
                    # [, ,赞]
                    elif index_3[0]<index_1[0]:
                        stop_index.append(index_3[0])
                        index_3.pop(0)    
                    else:
                        pass
                # index_1
                else:
                    # [分享, , ]
                    stop_index.append(index_1[0])
                    index_1.pop(0)
        else:
            if index_2:
                # index_2+index_3
                if index_3:
                    # ['', 评论, 赞]
                    if index_3[0]-index_2[0]==1:
                        stop_index.append(index_3[0])
                        index_2.pop(0)
                        index_3.pop(0)    
                    # [,评论 , ]
                    elif index_3[0]-index_2[0]>1:
                        stop_index.append(index_2[0])
                        index_2.pop(0)
                    # [, ,赞]
                    elif index_3[0]<index_2[0]:
                        stop_index.append(index_3[0])
                        index_3.pop(0)    
                    else:
                        pass
                # index_2
                else:
                    # [, 评论,]
                    stop_index.append(index_2[0])
                    index_2.pop(0)
            else:
                # index_3
                if index_3:
                    # [, ,赞]
                    stop_index.append(index_3[0])
                    index_3.pop(0)
                # none
                else:
                    pass
    # [问董秘, xx, 评论, ]
    if stop_index[0]-start_index<3:
        pass
    else:
        stop_index.insert(0,start_index)
    stop_index.append(end_index)
    result = []
    while len(stop_index)>1:
        temp = text[stop_index[0]+1:stop_index[1]+1]
        stop_index.pop(0)
        result.append(temp)
    return result

# write the review text to excel file 
# one review one line 
def write_xlxs(texts, filefolder, excel_name):
    excel_address = os.path.join(filefolder, excel_name)
    if not os.path.exists(excel_address):
        data = openpyxl.Workbook()
        table = data.create_sheet('Sheet1')
        table = data.active
        nrows = 1
    else:
        data = openpyxl.load_workbook(excel_address)
        # print(data.get_named_ranges()) # 输出工作页索引范围
        # print(data.get_sheet_names()) # 输出所有工作页的名称
        # 取第一张表
        sheetnames = data.get_sheet_names()
        table = data.get_sheet_by_name(sheetnames[0])
        table = data.active
        nrows = table.max_row # 获得行数
    # print(nrows)
    for i in range(len(texts)):
        text = texts[i]
        status = 0
        # status 0: 用户名+评论
        # status 1: 设备+评论
        # status 2: 用户名+设备+评论
        if 'iPhone' in text[0] or 'Android' in text[0] or 'PC' in text[0]:
            temp = text[0]
            status = 1
        if 'iPhone' in text[1] or 'Android' in text[1] or 'PC' in text[1]:
            temp = text[1]
            status = 2
        # 存储用户名        
        if status == 0 or status == 2:
            table.cell(row = int(nrows+i+1), column = 1).value = text.pop(0)
        # 存储设备+时间
        if status == 1 or status == 2:
            temp = text.pop(0)
            platform_time = temp.replace(" ","").split('发布于')
            platform = platform_time[0]
            time = platform_time[1].replace("-","")
            table.cell(row = int(nrows+i+1), column = 2).value = platform
            table.cell(row = int(nrows+i+1), column = 3).value = time
        # 存储评论
        index = []
        for j,v in enumerate(text):
            if ('分享' in v) or ('评论' in v) or ('赞'==v) or ('说两句' in v):
                index.append(j)
        end = min(index) if index else len(text)-1
        # print(text)
        for x in text[:end]:
            if len(x)<2:
                text.remove(x) 
        temp_text = ''.join(text[:end])
        table.cell(row = int(nrows+i+1), column = 4).value = temp_text
    data.save(excel_address)


def main():
    second = sleeptime(0,0,2)
    filefolder = r''
    excel_name = r'1.xlsx'
    texts = []
    text_count = 0
    for file in os.listdir(filefolder):
        # only process image file
        if os.path.splitext(file)[-1] in ['.png','.jpg','.jpeg','.bmp']:
            filePath = os.path.join(filefolder, file)
            time.sleep(second)
            raw_text = img_to_str(filePath)
            text = text_process(raw_text)  
            texts.extend(text)
            text_count += 1
            # process 100 pictures, write excel file one time
            if text_count > 99:
                write_xlxs(texts)
                text_count = 0
                texts = []
        else:
            pass
    print(texts)
    write_xlxs(texts,filefolder,excel_name)    

if __name__ == '__main__':
    main()
